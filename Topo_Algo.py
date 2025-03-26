import pandas as pd
import os
import glob
import logging
import os
import sys
import torch
import pandas as pd
import numpy as np
from collections import defaultdict
from torch.utils.data import Subset, DataLoader
from typing import OrderedDict, List, Optional
import networkx as nx
from sklearn.cluster import KMeans
from sklearn.mixture import GaussianMixture
from sklearn.metrics import confusion_matrix, precision_score, recall_score, f1_score, accuracy_score
from openpyxl import load_workbook
from openpyxl import Workbook
import re


class Topo_Algo:
    def __init__(self, case_dict, graph_object, excel_path, attack_path, round_case=None):
        self.case_dict = case_dict
        self.graph = graph_object
        self.excel_path = excel_path
        self.attack_path = attack_path
        self.round_case = round_case
    
    
    def get_actual_adj(self, G):
        adj_matrix = nx.adjacency_matrix(G).toarray()
    
        n = adj_matrix.shape[0]
        
        result = np.zeros((n, n-1), dtype=int)
    
        for i in range(n):
            result[i] = np.delete(adj_matrix[i], i) # Here, we try to remove the diagonal line elements to faciliate comparison
        
        return result
        
        
    def get_actual_adj_ls(self, G):
        adj_matrix = nx.adjacency_matrix(G).toarray()
        
        n = adj_matrix.shape[0]
    
        upper_triangle_list = []
        
        for i in range(n):
            for j in range(i + 1, n):
                upper_triangle_list.append(adj_matrix[i, j]) # Here, we only extract the upper triangle adj and store it as a list
                
        return upper_triangle_list
        
        
    def extract_data_from_directory(self):
        directory_path = self.excel_path
        
        file_pattern = os.path.join(directory_path, f"Round_{self.round_case-1}.xlsx")
        
        files = sorted(glob.glob(file_pattern))
        
        if not files:
          raise FileNotFoundError(f"No files found in the directory {directory_path} matching the pattern 'Round_*.xlsx'")
        
        df = pd.read_excel(files[0]) # here since we only extract from Round_9, so the files is single.
        
        return df
        
        
    def create_workbook(self, file_name):
        """Create a new Excel workbook with headers."""
        wb = Workbook()
        ws = wb.active
        # Add headers to the first row if necessary
        if "Supervised" not in file_name:
            ws.append(["File Name", "Metric", "Algo", "Clustering", "Accuracy", "Precison", "FPR", "F1", "Actual_Adj", "Inferred_Adj"])
        else:
            ws.append(["File Name", "Metric", "Model", "Algo", "Ratio", "Train_Size", "Test_Size", "Accuracy", "Macro_precision", "Macro_recall", 
                       "Macro_F1", "Weighted_precision","Weighted_recall", "Weighted_F1"])
                       
            
        wb.save(file_name)
    
    
    def append_experiment_results(self, file_name, data):
        # Check if the file exists
        if not os.path.exists(file_name):
            # If the file does not exist, create it
            self.create_workbook(file_name)
        
        # Now load the workbook and append data
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append(data)
        wb.save(file_name)
    
    
    def evaluate_inference(self, inferred_adj, actual_adj):
        actual_adj = actual_adj.flatten() if not isinstance(actual_adj, list) else actual_adj
        inferred_adj = inferred_adj.flatten() if not isinstance(inferred_adj, list) else inferred_adj
    
        tn, fp, fn, tp = confusion_matrix(actual_adj, inferred_adj).ravel()
        '''print(tn)
        print(fp)
        print(fn)
        print(tp)'''
        tpr = tp / (tp + fn)  # True Positive Rate
        fpr = fp / (fp + tn)  # False Positive Rate
        precision = precision_score(actual_adj, inferred_adj)
        f1 = f1_score(actual_adj, inferred_adj)
        accuracy = accuracy_score(actual_adj, inferred_adj)
        
        return accuracy, tpr, fpr, f1
        
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
        
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
        
    
    